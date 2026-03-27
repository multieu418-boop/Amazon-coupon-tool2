import streamlit as st
import pandas as pd
import re
import math
import io
import openpyxl
from copy import copy

st.set_page_config(page_title="Amazon Coupon 专家级修复工具", layout="wide")

# --- 1. 核心解析逻辑：精准提取报错原因和价格 ---
def parse_error_details(comment_text):
    error_map = {}
    if not comment_text: return error_map
    # 匹配 10位ASIN + 换行
    blocks = re.split(r'([A-Z0-9]{10})\n', str(comment_text))
    if len(blocks) > 1:
        for i in range(1, len(blocks), 2):
            asin = blocks[i].strip()
            content = blocks[i+1]
            
            # 兼容多种亚马逊报错价格描述
            req_p_match = re.search(r'(?:要求的净价格|当前净价格|要求的最高商品价格)：[^\d]*([\d\.]+)', content)
            req_p = float(req_p_match.group(1)) if req_p_match else None
            
            # 提取原因：ASIN 换行后到“价格”字样前的所有描述文字
            reason_part = re.split(r'(?:要求的净价格|当前净价格|要求的最高商品价格)', content)[0]
            reason = reason_part.strip().replace('\n', ' ')
            
            error_map[asin] = {"req_price": req_p, "reason": reason}
    return error_map

# --- 2. 底稿读取：保留所有原表格式 ---
def load_excel_template(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    data = []
    headers = [cell.value for cell in ws[7]] # 亚马逊模板第7行通常是表头
    for row_idx, row in enumerate(ws.iter_rows(min_row=10), 10):
        row_values = [cell.value for cell in row]
        if not any(row_values): continue
        # 最后一列通常是批注列
        comment = row[-1].comment.text if row[-1].comment else ""
        row_dict = {headers[i]: val for i, val in enumerate(row_values) if i < len(headers)}
        row_dict['_comment'] = comment
        row_dict['_row_idx'] = row_idx
        data.append(row_dict)
    return pd.DataFrame(data), headers

# --- 3. Listing 报告读取 ---
def load_listing(file):
    for enc in ['utf-8', 'utf-16', 'gbk', 'utf-8-sig']:
        try:
            file.seek(0)
            df = pd.read_csv(file, sep='\t', encoding=enc) if file.name.endswith('.txt') else pd.read_excel(file)
            df.columns = [c.lower().strip() for c in df.columns]
            return df
        except: continue
    return None

def main():
    st.title("🎯 Amazon Coupon 终极全量修复系统")
    
    # 初始化全局 Master Data，防止筛选时数据重置
    if 'master_df' not in st.session_state:
        st.session_state.master_df = None
        st.session_state.orig_headers = None

    # --- 侧边栏：需求全量保留 ---
    st.sidebar.header("⚙️ 筛选与预警配置")
    status_sel = st.sidebar.multiselect("1. ASIN 状态筛选", ["✅ 正常", "❌ 批注报错"], default=["✅ 正常", "❌ 批注报错"])
    discount_threshold = st.sidebar.slider("2. 折扣力度红色预警线 (%)", 5, 50, 30) / 100
    reason_keyword = st.sidebar.text_input("3. 报错原因关键词过滤")
    
    if st.sidebar.button("🔄 重置/重新上传"):
        st.session_state.master_df = None
        st.rerun()

    col1, col2 = st.columns(2)
    with col1:
        l_file = st.file_uploader("1. 上传 All Listing 报告", type=['txt', 'xlsx', 'csv'])
    with col2:
        e_file = st.file_uploader("2. 上传带批注的报错模板", type=['xlsx'])

    if l_file and e_file:
        if st.session_state.master_df is None:
            df_l = load_listing(l_file)
            e_file.seek(0)
            df_ui, headers = load_excel_template(e_file)
            
            if df_l is not None and df_ui is not None:
                # 动态定位关键列
                asin_col = next((c for c in df_l.columns if 'asin' in c), None)
                price_col = next((c for c in df_l.columns if 'price' in c or '价格' in c), None)
                e_asin_idx = next((c for c in df_ui.columns if 'ASIN' in str(c)), df_ui.columns[0])
                e_disc_idx = next((c for c in df_ui.columns if '折扣' in str(c) and '数值' in str(c)), None)

                all_rows = []
                for _, row in df_ui.iterrows():
                    asins = [a.strip() for a in str(row.get(e_asin_idx, "")).replace(',', ';').split(';') if a.strip()]
                    err_map = parse_error_details(row.get('_comment'))
                    
                    for a in asins:
                        p_match = df_l[df_l[asin_col] == a][price_col].values if asin_col else []
                        orig_p = p_match[0] if len(p_match) > 0 else None
                        info = err_map.get(a, {})
                        is_err = a in err_map
                        
                        # 初始建议折扣
                        curr_d = row.get(e_disc_idx, 0.05)
                        suggested = curr_d
                        if is_err and orig_p and info.get('req_price'):
                            needed = math.ceil(((float(orig_p) - float(info.get('req_price'))) / float(orig_p)) * 100)
                            suggested = needed / 100 if curr_d < 1 else needed

                        all_rows.append({
                            "决策": "保留",
                            "ASIN": a,
                            "状态": "❌ 批注报错" if is_err else "✅ 正常",
                            "详细报错原因": info.get('reason', "-"),
                            "拟提报折扣": suggested,
                            "Listing原价": orig_p,
                            "要求净价": info.get('req_price'),
                            "原始行号": row.get('_row_idx'),
                            "meta_row": row.to_dict()
                        })
                st.session_state.master_df = pd.DataFrame(all_rows)
                st.session_state.orig_headers = headers

        if st.session_state.master_df is not None:
            # 应用侧边栏筛选逻辑
            mask = st.session_state.master_df['状态'].isin(status_sel)
            if reason_keyword:
                mask = mask & st.session_state.master_df['详细报错原因'].str.contains(reason_keyword, case=False)
            
            df_filtered = st.session_state.master_df[mask].copy()

            st.subheader("🛠️ 修复决策台")
            
            # 样式：高折扣预警渲染（仅用于预览）
            def style_discount(v):
                return 'color: red; font-weight: bold' if v > discount_threshold else ''

            # 使用 data_editor
            edited_data = st.data_editor(
                df_filtered,
                column_config={
                    "决策": st.column_config.SelectboxColumn("决策", options=["保留", "剔除"]),
                    "拟提报折扣": st.column_config.NumberColumn("拟提报折扣", format="%.2f"),
                    "详细报错原因": st.column_config.TextColumn("详细报错原因", width="large"),
                    "meta_row": None, "原始行号": None
                },
                disabled=['ASIN', '状态', '详细报错原因', 'Listing原价', '要求净价'],
                hide_index=True,
                use_container_width=True,
                key="editor_main"
            )

            # --- 关键修复：即时同步修改到总表 ---
            if not edited_data.equals(df_filtered):
                # 找出发生变化的行并同步回 master_df
                # 我们使用原始 dataframe 的 index 作为对照
                for idx in edited_data.index:
                    st.session_state.master_df.loc[idx, '决策'] = edited_data.loc[idx, '决策']
                    st.session_state.master_df.loc[idx, '拟提报折扣'] = edited_data.loc[idx, '拟提报折扣']
                st.rerun() # 强制刷新页面以固化状态

            # --- 4. 导出逻辑：全量无损合并 ---
            st.markdown("---")
            if st.button("🚀 导出全量原格式文件 (包含正确 + 已保留报错项)"):
                e_file.seek(0)
                wb = openpyxl.load_workbook(e_file)
                ws = wb.active
                
                # 清空底稿第10行以后所有数据
                for r in range(10, ws.max_row + 1):
                    ws.cell(row=r, column=1).value = None 

                # 从总表中获取所有“保留”的项，不论当前页面筛选了什么
                final_keep = st.session_state.master_df[st.session_state.master_df['决策'] == "保留"]
                
                if final_keep.empty:
                    st.warning("无可导出的保留项，请检查决策列。")
                else:
                    a_idx, d_idx = 1, 3
                    for i, h in enumerate(st.session_state.orig_headers, 1):
                        if h and 'ASIN' in str(h): a_idx = i
                        if h and '折扣' in str(h) and '数值' in str(h): d_idx = i

                    curr_r = 10
                    # 按原始行号和折扣分组（剥离逻辑）
                    for (orig_line, disc), group in final_keep.groupby(['原始行号', '拟提报折扣']):
                        # 复制原行格式和所有列数据
                        for c in range(1, len(st.session_state.orig_headers) + 1):
                            source = ws.cell(row=orig_line, column=c)
                            target = ws.cell(row=curr_r, column=c)
                            target.value = source.value
                            # 无损复制样式
                            if source.has_style:
                                target.font = copy(source.font)
                                target.border = copy(source.border)
                                target.fill = copy(source.fill)
                                target.number_format = copy(source.number_format)
                                target.alignment = copy(source.alignment)
                        
                        # 覆盖当前行的 ASIN 串和折扣
                        ws.cell(row=curr_r, column=a_idx).value = ";".join(group['ASIN'].tolist())
                        ws.cell(row=curr_r, column=d_idx).value = disc
                        curr_r += 1

                    # 删除底稿多余的行
                    if ws.max_row >= curr_r:
                        ws.delete_rows(curr_r, ws.max_row - curr_r + 1)
                    
                    out_io = io.BytesIO()
                    wb.save(out_io)
                    st.success(f"导出成功！文件包含 {len(final_keep)} 个 ASIN。")
                    st.download_button("📥 点击下载最终修复文件", out_io.getvalue(), "Amazon_Fixed_All_ASINs.xlsx")

if __name__ == "__main__":
    main()
